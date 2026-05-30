"""
知乎文章/回答批量爬虫 — Scrapling 版

核心策略：
  知乎的 x-zse-96 / x-zst-81 由页面 JS 动态生成，且依赖 canvas 指纹等
  浏览器环境。与其逆向加密算法（极易因版本更新失效），不如利用 Scrapling
  的 DynamicSession 在真实浏览器上下文中调用 fetch()，让知乎自己的
  fetch hook 自动附上签名头。

  相比纯 Playwright 版，本版额外享受：
  - Scrapling 内置的反检测（block ads、dns over https、指纹伪装）
  - Scrapling 的 CSS/XPath 解析引擎（用于 HTML 页面提取）
  - Scrapling 的 ProxyRotator（代理池轮换）
  - Scrapling 的 Spider 分页框架

使用：
  pip install "scrapling[fetchers]"
  scrapling install
  python zhihu_scrapling.py user <用户ID>
"""

from __future__ import annotations

import json
import os
import re
import time
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from scrapling.fetchers import DynamicSession

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("zhihu_scrapling")

# ── 常量 ──────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent / "output"
DEFAULT_LIMIT = 20
REQUEST_DELAY = 1.5
PAGE_TIMEOUT = 30_000


class ZhihuScraplingSpider:
    """
    基于 Scrapling DynamicSession 的知乎爬虫。

    工作原理：
      1. DynamicSession 启动 Chromium，打开知乎首页
      2. 知乎 JS 加载完成，hook 了 window.fetch
      3. 通过 page.evaluate() 在浏览器上下文调 fetch()，
         x-zse-96 / x-zst-81 自动由知乎的 hook 附加
      4. 返回的 JSON 数据经由 Python 处理并持久化
    """

    def __init__(self, headless: bool = False, profile_dir: Optional[Path] = None):
        self.headless = headless
        self.profile_dir = str(profile_dir or (Path(__file__).parent / "zhihu_profile"))
        self._session: Optional[DynamicSession] = None
        self._playwright_page = None

    # ── 生命周期 ────────────────────────────────────
    def __enter__(self):
        self.start()
        return self

    def __exit__(self, *_):
        self.close()

    def start(self):
        """启动浏览器会话，打开知乎并等待 JS 就绪。"""
        logger.info("启动 Scrapling DynamicSession...")
        os.makedirs(self.profile_dir, exist_ok=True)

        self._session = DynamicSession(
            headless=self.headless,
            network_idle=True,
            timeout=PAGE_TIMEOUT,
            disable_resources=False,
            block_ads=True,
            dns_over_https=True,
            user_data_dir=self.profile_dir,
        )

        self._session.start()

        logger.info("打开知乎首页...")
        # fetch() 用 page pool 打开页面加载 JS 钩子，用完页面可能回收
        resp = self._session.fetch(
            "https://www.zhihu.com",
            timeout=PAGE_TIMEOUT,
            network_idle=True,
            wait=3000,
        )
        logger.info(f"知乎首页状态: {resp.status}")

        # 从 persistent context 拿一个持久化的 page 用于后续 API 调用
        self._playwright_page = self._acquire_page()
        self._playwright_page.goto("https://www.zhihu.com", timeout=PAGE_TIMEOUT)
        self._playwright_page.wait_for_load_state("networkidle")
        time.sleep(2)

        if not self._check_login():
            logger.warning("未检测到登录状态，请在浏览器窗口中扫码登录")
            logger.warning("登录后按 Enter 继续...")
            input()
            if not self._check_login():
                logger.warning("仍未检测到登录状态，继续尝试...")
        else:
            logger.info("已登录 ✓")

    def close(self):
        if self._session:
            try:
                self._session.close()
            except Exception:
                pass

    # ── 内部工具 ────────────────────────────────────
    def _acquire_page(self):
        """从 session.context 获取或创建一个 page。"""
        ctx = self._session.context
        if ctx.pages:
            return ctx.pages[0]
        return ctx.new_page()

    def _ensure_page(self):
        """确保 Playwright Page 可用。"""
        if self._playwright_page is None:
            self._playwright_page = self._acquire_page()
        try:
            self._playwright_page.title()
        except Exception:
            logger.info("页面已关闭，重新打开...")
            self._playwright_page = self._session.context.new_page()
            self._playwright_page.goto("https://www.zhihu.com", timeout=PAGE_TIMEOUT)
            self._playwright_page.wait_for_load_state("networkidle")
            time.sleep(2)

    def _check_login(self) -> bool:
        try:
            result = self._playwright_page.evaluate("""() => {
                return document.cookie.includes('z_c0=');
            }""")
            return bool(result)
        except Exception:
            return False

    # ── 核心：浏览器内 API 调用 ─────────────────────
    def api_call(self, path: str, params: Optional[dict] = None,
                 method: str = "GET") -> dict:
        """
        在浏览器上下文中调用知乎 API。
        知乎的 fetch hook 自动附加 x-zse-96 / x-zst-81，无需手动逆向。
        """
        self._ensure_page()

        if params is None:
            params = {}
        full_url = f"https://www.zhihu.com{path}"
        if params:
            full_url += "?" + urlencode(params, doseq=True)

        js_code = """
            async ([fullUrl, method]) => {
                const resp = await fetch(fullUrl, {
                    method: method,
                    credentials: 'include',
                    headers: {
                        'x-requested-with': 'fetch',
                        'content-type': 'application/json',
                    },
                });
                const text = await resp.text();
                try {
                    return JSON.parse(text);
                } catch {
                    return { _raw: text, _status: resp.status, _ok: resp.ok };
                }
            }
        """

        resp = self._playwright_page.evaluate(js_code, [full_url, method])
        time.sleep(REQUEST_DELAY)
        return resp

    def api_paginate(self, path: str, params: dict,
                     max_items: int = 0,
                     list_key: str = "data",
                     offset_key: str = "offset",
                     limit_key: str = "limit") -> list[dict]:
        """通用分页器 — 自动翻页直到取完或达到上限。"""
        params = dict(params)
        params.setdefault(limit_key, DEFAULT_LIMIT)
        all_items = []
        offset = params.get(offset_key, 0)

        while True:
            params[offset_key] = offset
            resp = self.api_call(path, params)

            if not resp or resp.get(list_key) is None:
                logger.warning(f"分页中断: {str(resp)[:200]}")
                break

            page_data = resp[list_key]
            if not page_data:
                break

            all_items.extend(page_data)
            logger.info(f"  已获取 {len(all_items)} 条...")

            if max_items and len(all_items) >= max_items:
                return all_items[:max_items]

            if resp.get("paging", {}).get("is_end"):
                break

            offset += len(page_data)

        return all_items

    # ── 业务 API ─────────────────────────────────────
    def get_user_info(self, user_id: str) -> dict:
        include = "answer_count,articles_count,follower_count,voteup_count,gender,headline,avatar_url"
        return self.api_call(f"/api/v4/members/{user_id}", {"include": include})

    def get_user_answers(self, user_id: str, max_count: int = 0,
                         sort_by: str = "created") -> list[dict]:
        logger.info(f"获取用户 {user_id} 的回答...")
        return self.api_paginate(
            f"/api/v4/members/{user_id}/answers",
            {
                "include": (
                    "data[*].is_normal,content,excerpt,voteup_count,created_time,"
                    "updated_time,comment_count,question,relationship.is_authorized"
                ),
                "limit": DEFAULT_LIMIT,
                "offset": 0,
                "sort_by": sort_by,
            },
            max_items=max_count,
        )

    def get_user_articles(self, user_id: str, max_count: int = 0) -> list[dict]:
        logger.info(f"获取用户 {user_id} 的文章...")

        cols = self.api_call(
            f"/api/v4/members/{user_id}/column-contributions",
            {"include": "column", "limit": 50, "offset": 0},
        )
        columns = []
        for item in (cols.get("data") or []):
            if "column" in item:
                columns.append(item["column"])

        if not columns:
            logger.warning("未找到专栏")
            return []

        all_articles = []
        for col in columns:
            logger.info(f"  专栏: {col.get('title', col['id'])}")
            articles = self.api_paginate(
                f"/api/v4/columns/{col['id']}/articles",
                {
                    "include": "data[*].title,content,excerpt,voteup_count,"
                               "created_time,updated_time,comment_count,image_url",
                    "limit": DEFAULT_LIMIT,
                    "offset": 0,
                },
                max_items=max_count,
            )
            all_articles.extend(articles)

        logger.info(f"文章获取完成，共 {len(all_articles)} 篇")
        return all_articles

    def get_question_answers(self, question_id, max_count: int = 0,
                             order: str = "default") -> list[dict]:
        logger.info(f"获取问题 {question_id} 的回答...")
        return self.api_paginate(
            f"/api/v4/questions/{question_id}/answers",
            {
                "include": (
                    "data[*].is_normal,content,excerpt,voteup_count,created_time,"
                    "updated_time,comment_count;data[*].author.follower_count,"
                    "badge[*].topics"
                ),
                "limit": DEFAULT_LIMIT,
                "offset": 0,
                "order": order,
            },
            max_items=max_count,
        )

    def search(self, query: str, search_type: str = "content",
               max_count: int = 0) -> list[dict]:
        logger.info(f"搜索: {query}")
        return self.api_paginate(
            "/api/v4/search_v3",
            {
                "t": search_type,
                "q": query,
                "correction": 1,
                "offset": 0,
                "limit": DEFAULT_LIMIT,
                "lc_idx": 0,
                "show_all_topics": 0,
            },
            max_items=max_count,
        )

    def get_related_questions(self, question_id, max_count: int = 0) -> list[dict]:
        """获取相关问题推荐。"""
        return self.api_paginate(
            f"/api/v4/questions/{question_id}/related-questions",
            {"limit": DEFAULT_LIMIT, "offset": 0},
            max_items=max_count,
        )

    # ── 数据导出 ────────────────────────────────────
    @staticmethod
    def _strip_html(text: str) -> str:
        """去掉 HTML 标签，保留纯文本。"""
        if not text:
            return ""
        return re.sub(r"<[^>]+>", "", text).strip()

    @staticmethod
    def _ts_to_str(ts) -> str:
        """Unix 时间戳 → 可读字符串。"""
        if not ts:
            return ""
        try:
            return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        except (ValueError, OSError):
            return str(ts)

    @staticmethod
    def _auto_width(ws, min_width: int = 8, max_width: int = 60):
        """根据内容自适应列宽。"""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    # 中文字符算 2 宽
                    val = str(cell.value)
                    length = sum(2 if ord(ch) > 127 else 1 for ch in val)
                    max_len = max(max_len, length)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    def save_excel(
        self,
        answers: list[dict] | None = None,
        articles: list[dict] | None = None,
        filename: str = "zhihu_export.xlsx",
    ):
        """将回答和文章写入 Excel 文件，每个类型一个 sheet。"""
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        path = OUTPUT_DIR / filename

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 移除默认空 sheet

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="top", wrap_text=True)

        def write_sheet(ws, headers: list[str], rows: list[list]):
            """写入表头和数据，设置样式。"""
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = cell_align
            ws.freeze_panes = "A2"
            self._auto_width(ws)

        # ── 回答 Sheet ──
        if answers:
            ans_headers = [
                "序号", "问题", "回答摘要", "赞同数", "评论数",
                "创建时间", "更新时间", "链接",
            ]
            ans_rows = []
            for i, a in enumerate(answers, 1):
                q = a.get("question", {}) or {}
                ans_rows.append([
                    i,
                    self._strip_html(q.get("title", "")),
                    self._strip_html(a.get("excerpt", "")),
                    a.get("voteup_count", 0),
                    a.get("comment_count", 0),
                    self._ts_to_str(a.get("created_time")),
                    self._ts_to_str(a.get("updated_time")),
                    f"https://www.zhihu.com/question/{q.get('id', '')}/answer/{a.get('id', '')}",
                ])
            ws_ans = wb.create_sheet("回答")
            write_sheet(ws_ans, ans_headers, ans_rows)

        # ── 文章 Sheet ──
        if articles:
            art_headers = [
                "序号", "标题", "摘要", "赞同数", "评论数",
                "创建时间", "更新时间", "链接",
            ]
            art_rows = []
            for i, a in enumerate(articles, 1):
                art_rows.append([
                    i,
                    self._strip_html(a.get("title", "")),
                    self._strip_html(a.get("excerpt", "")),
                    a.get("voteup_count", 0),
                    a.get("comment_count", 0),
                    self._ts_to_str(a.get("created_time")),
                    self._ts_to_str(a.get("updated_time")),
                    a.get("url", f"https://zhuanlan.zhihu.com/p/{a.get('id', '')}"),
                ])
            ws_art = wb.create_sheet("文章")
            write_sheet(ws_art, art_headers, art_rows)

        wb.save(path)
        logger.info(f"已保存 Excel → {path}")

    def save_json(self, data, filename: str):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        path = OUTPUT_DIR / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"已保存 → {path}")


# ══════════════════════════════════════════════════════
# 便捷入口
# ══════════════════════════════════════════════════════

def crawl_user(user_id: str, headless: bool = False):
    with ZhihuScraplingSpider(headless=headless) as spider:
        info = spider.get_user_info(user_id)
        logger.info(f"用户: {info.get('name', user_id)}, "
                     f"回答 {info.get('answer_count', '?')}, "
                     f"文章 {info.get('articles_count', '?')}")

        answers = spider.get_user_answers(user_id)
        articles = spider.get_user_articles(user_id)

        spider.save_json(answers, f"{user_id}_answers.json")
        spider.save_json(articles, f"{user_id}_articles.json")
        spider.save_json({
            "user_info": info,
            "answers": answers,
            "articles": articles,
        }, f"{user_id}_all.json")
        spider.save_excel(answers=answers, articles=articles,
                          filename=f"{user_id}_all.xlsx")

        return answers, articles


def crawl_question(question_id, headless: bool = False):
    with ZhihuScraplingSpider(headless=headless) as spider:
        answers = spider.get_question_answers(question_id)
        spider.save_json(answers, f"question_{question_id}_answers.json")
        spider.save_excel(answers=answers,
                          filename=f"question_{question_id}.xlsx")
        return answers


# ══════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="知乎爬虫 (Scrapling)")
    sp = p.add_subparsers(dest="cmd")

    up = sp.add_parser("user", help="爬取用户回答+文章")
    up.add_argument("user_id")
    up.add_argument("--headless", action="store_true")

    qp = sp.add_parser("question", help="爬取问题回答")
    qp.add_argument("question_id")
    qp.add_argument("--headless", action="store_true")

    args = p.parse_args()
    if args.cmd == "user":
        crawl_user(args.user_id, headless=args.headless)
    elif args.cmd == "question":
        crawl_question(args.question_id, headless=args.headless)
    else:
        p.print_help()
